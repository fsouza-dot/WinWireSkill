#!/usr/bin/env python3
"""Read a simplified WinWire template (.xlsx) and convert it to the JSON data
format expected by build_html.py.

The simplified template only has the fields that humans uniquely know (deal metrics,
project identity, challenge/solution highlights). Everything else — narrative polish,
technologies, page 2 phases, tech architecture, quotes — is expected to come from
attached project documents, handled by the agentic skill layer.

This script focuses on extracting what's in the template. The skill orchestrates
the full pipeline: template fields + doc extraction + user Q&A for gaps.

Usage:
    python ingest.py --template simplified-template.xlsx --output project-data.json
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(1)


def read_template(template_path: Path) -> dict:
    """Read the xlsx template and extract all fields into a dict keyed by field label."""
    wb = openpyxl.load_workbook(str(template_path), data_only=True)
    ws = wb.active

    fields = {}
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if label:
            # Strip the required marker and clean up
            clean_label = str(label).strip().replace(" *", "")
            raw_value = str(value).strip() if value else ""
            # Skip placeholder/example text (italic helper text starts with "e.g." or "Type" or "Leave blank")
            if raw_value.startswith("e.g.,") or raw_value.startswith("e.g. ") or \
               raw_value.startswith("Type ") or raw_value.startswith("Leave blank") or \
               raw_value.startswith("aws, gcp, or azure"):
                raw_value = ""
            fields[clean_label] = raw_value

    return fields


def parse_comma_list(text):
    """Split a comma-separated string into a list of stripped strings."""
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def fields_to_json(fields: dict) -> dict:
    """Convert flat field dict to the nested JSON structure for build_html.py.

    Returns a partial JSON — fields that the skill needs to fill from docs
    will be empty strings or empty lists. The skill layer is responsible for
    merging doc-extracted content into these gaps.
    """
    def get(key, default=""):
        return fields.get(key, default) or default

    anonymize = get("Anonymize?").upper() in ("YES", "TRUE", "Y", "1")

    # Parse quote author + title if provided as combined field
    quote_author = ""
    quote_title = ""
    quote_author_raw = get("Quote Author + Title")
    if quote_author_raw and "," in quote_author_raw:
        parts = quote_author_raw.split(",", 1)
        quote_author = parts[0].strip()
        quote_title = parts[1].strip()
    elif quote_author_raw:
        quote_author = quote_author_raw

    data = {
        "project": {
            "client_name": get("Client Name"),
            "anonymized_name": get("Anonymized Name"),
            "anonymize": anonymize,
            "industry": get("Industry"),
            "partner": get("Cloud Partner").lower() if get("Cloud Partner") else "",
            "project_type": get("Project Type"),
            # Title and subtitle may come from template or be drafted by the skill
            "title": get("Preferred Title"),
            "subtitle": get("Preferred Subtitle"),
            "tags": [],  # Skill infers from project type, industry, partner, technologies
        },

        "challenge": {
            "headline": "",  # Skill drafts from user highlights + docs
            "body": "",      # Skill drafts from user highlights + docs
            "_user_highlights": get("What was the challenge?"),  # Raw user input
        },

        "solution": {
            "headline": "",      # Skill drafts from user highlights + docs
            "body": "",          # Skill drafts from user highlights + docs
            "technologies": [],  # Skill extracts from docs
            "_user_highlights": get("What did CI&T deliver?"),  # Raw user input
        },

        "metrics_internal": {
            "items": [
                {"value": get("Services Revenue"), "label": "Services Revenue"},
                {"value": get("Incentive Funding"), "label": "Incentive Funding"},
                {"value": get("Deal Cycle"), "label": "Deal Cycle"},
            ]
        },

        "quote": {
            "text": get("Client Quote"),
            "author": quote_author,
            "title": quote_title,
            "company": get("Client Name"),
        },

        "context_bar_internal": [],
        "context_bar_partner": [],

        "logos": {
            "cit_logo_url": "",
            "partner_logo_url": "",
            # Either a URL, a data: URL, or an absolute/relative local path. See SKILL.md
            # "Client logo resolution flow" for how the skill populates this.
            "client_logo_url": get("Client Logo URL"),
            "client_logo_file": get("Client Logo File"),
        },

        # Metadata for the skill layer — what still needs to be filled
        "_extraction_needed": [],
    }

    # Partner metrics — ACR is mandatory for the partner version (reserved slot #1).
    # We always emit the metrics_partner block so the reserved slot is visible downstream,
    # even if ACR is empty. The skill layer is responsible for recovering ACR from attached
    # docs, asking the user, or warning them before publishing without it.
    acr = get("Annual Cloud Revenue (ACR)")
    outcome = get("Key Business Outcome")
    services = get("Services Adopted")
    data["metrics_partner"] = {
        "items": [
            {"value": acr, "label": "Annual Cloud Revenue"},
            {"value": outcome, "label": "Infrastructure Cost Reduction"},
            {"value": services, "label": "Core Services Adopted"},
        ]
    }

    # Context bar — internal
    ctx_internal = []
    incentive_prog = get("Incentive Program")
    funding = get("Services Revenue")  # Reuse if no separate funding field
    cycle = get("Deal Cycle")
    team_size = get("Delivery Team Size")
    timeline = get("Delivery Timeline")
    incentive_funding = get("Incentive Funding")

    if incentive_prog:
        ctx_internal.append({"label": "Incentive Program", "value": incentive_prog})
    if incentive_funding:
        ctx_internal.append({"label": "Funding Received", "value": incentive_funding})
    if cycle:
        ctx_internal.append({"label": "Deal Cycle", "value": cycle})
    if team_size:
        ctx_internal.append({"label": "Delivery Team", "value": team_size})
    data["context_bar_internal"] = ctx_internal

    # Context bar — partner
    ctx_partner = []
    if incentive_prog:
        ctx_partner.append({"label": "Incentive Program", "value": incentive_prog})
    if incentive_funding:
        ctx_partner.append({"label": "Funding Received", "value": incentive_funding})
    if cycle:
        ctx_partner.append({"label": "Deal Cycle", "value": cycle})
    if timeline:
        ctx_partner.append({"label": "Delivery Timeline", "value": timeline})
    data["context_bar_partner"] = ctx_partner

    # Page 2 — always extracted from docs by the skill layer
    data["page2"] = {
        "include": False,  # Skill sets to True if enough content is found
        "deep_dive_title": "",
        "approach_title": "Project Approach",
        "phases": [],
        "metrics_table": [],
        "tech_architecture": [],
    }

    # Track what the skill needs to fill. Revenue gaps are flagged as CRITICAL because the
    # skill is required to resolve them (mine docs → ask user → warn) before publishing.
    gaps = []
    critical_gaps = []

    # Revenue metrics — reserved slot #1 on each version
    if not data["metrics_internal"]["items"][0]["value"]:
        critical_gaps.append("metrics_internal.services_revenue")
    if not data["metrics_partner"]["items"][0]["value"]:
        critical_gaps.append("metrics_partner.annual_cloud_revenue")

    if not data["project"]["title"]:
        gaps.append("project.title")
    if not data["project"]["subtitle"]:
        gaps.append("project.subtitle")
    if not data["challenge"]["_user_highlights"]:
        gaps.append("challenge.highlights")
    if not data["solution"]["_user_highlights"]:
        gaps.append("solution.highlights")
    if not data["quote"]["text"]:
        gaps.append("quote")
    # Non-revenue partner metrics
    if not data["metrics_partner"]["items"][1]["value"]:
        gaps.append("metrics_partner.business_outcome")
    if not data["metrics_partner"]["items"][2]["value"]:
        gaps.append("metrics_partner.services_adopted")
    # Page 2 always needs extraction
    gaps.append("page2.phases")
    gaps.append("page2.metrics_table")
    gaps.append("page2.tech_architecture")
    gaps.append("solution.technologies")
    gaps.append("project.tags")

    data["_extraction_needed"] = gaps
    data["_critical_gaps"] = critical_gaps

    return data


def validate(data: dict) -> tuple:
    """Check required fields. Returns (errors, warnings)."""
    errors = []
    warnings = []

    # Only hard-require identity/cycle/funding. Revenue metrics produce CRITICAL warnings
    # (not errors) because the skill is expected to recover them from docs or user Q&A
    # before build_html.py runs. Hard-failing here would block that recovery flow.
    required = [
        ("Client Name", data["project"]["client_name"]),
        ("Industry", data["project"]["industry"]),
        ("Cloud Partner", data["project"]["partner"]),
        ("Project Type", data["project"]["project_type"]),
        ("Incentive Funding", data["metrics_internal"]["items"][1]["value"]),
        ("Deal Cycle", data["metrics_internal"]["items"][2]["value"]),
    ]
    for field_name, value in required:
        if not value:
            errors.append(f"Missing required field: {field_name}")

    # CRITICAL: reserved revenue slots. The skill MUST mine docs / ask the user / warn
    # before publishing without these. See SKILL.md "Non-negotiable rule: the revenue metric".
    if not data["metrics_internal"]["items"][0]["value"]:
        warnings.append(
            "CRITICAL: Services Revenue (internal reserved slot #1) is missing — "
            "skill must mine docs or ask user; warn user before publishing without it."
        )
    if not data["metrics_partner"]["items"][0]["value"]:
        warnings.append(
            "CRITICAL: Annual Cloud Revenue / ACR (partner reserved slot #1) is missing — "
            "skill must mine docs or ask user; warn user before publishing without it."
        )

    # Warnings for things the skill will need to handle
    if not data["challenge"].get("_user_highlights"):
        warnings.append("No challenge highlights provided — skill will need docs or follow-up questions")
    if not data["solution"].get("_user_highlights"):
        warnings.append("No solution highlights provided — skill will need docs or follow-up questions")
    if not data["quote"]["text"]:
        warnings.append("No client quote — skill will ask the user")
    if not data["metrics_partner"]["items"][1]["value"] and not data["metrics_partner"]["items"][2]["value"]:
        warnings.append("No partner business-outcome or services-adopted numbers — partner version will need these")

    return errors, warnings


def main():
    parser = argparse.ArgumentParser(description="Convert simplified WinWire template (.xlsx) to JSON")
    parser.add_argument("--template", required=True, help="Path to the filled-in .xlsx template")
    parser.add_argument("--output", required=True, help="Output JSON file path")
    args = parser.parse_args()

    template_path = Path(args.template)
    if not template_path.exists():
        print(f"Error: template file not found: {template_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading template: {template_path}")
    fields = read_template(template_path)

    filled = sum(1 for v in fields.values() if v)
    print(f"Fields with data: {filled}/{len(fields)}")

    data = fields_to_json(fields)

    errors, warnings = validate(data)
    if errors:
        print("\nErrors (must fix):")
        for e in errors:
            print(f"  ✗ {e}")
    if warnings:
        print("\nInfo (skill will handle):")
        for w in warnings:
            print(f"  → {w}")

    gaps = data.get("_extraction_needed", [])
    if gaps:
        print(f"\nFields for doc extraction / user Q&A: {', '.join(gaps)}")

    # Write JSON
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"\nJSON written to: {output_path}")
    if errors:
        print("⚠ Fix the errors above before proceeding.")
        sys.exit(1)


if __name__ == "__main__":
    main()

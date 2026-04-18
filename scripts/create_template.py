#!/usr/bin/env python3
"""Generate a simplified WinWire input template as an .xlsx file.

The simplified template only asks for what a human uniquely knows: project identity,
deal metrics, and free-text challenge/solution highlights. Everything else (narratives,
tech details, page 2, quotes) is extracted from attached project documents by the skill.

Usage:
    python create_template.py --output /path/to/winwire-template.xlsx
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl --break-system-packages",
          file=sys.stderr)
    sys.exit(1)

CORAL = "FA5A50"
NAVY = "000050"
WHITE = "FFFFFF"
MID_GRAY = "DCE8F5"
LIGHT_BG = "F8F9FC"


def create_template(output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WinWire Input"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 65
    ws.sheet_properties.tabColor = CORAL

    # Styles
    header_font = Font(name="Arial", size=14, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=CORAL, end_color=CORAL, fill_type="solid")
    section_font = Font(name="Arial", size=12, bold=True, color=NAVY)
    section_fill = PatternFill(start_color=MID_GRAY, end_color=MID_GRAY, fill_type="solid")
    label_font = Font(name="Arial", size=11, bold=True, color=NAVY)
    req_label_font = Font(name="Arial", size=11, bold=True, color=CORAL)
    helper_font = Font(name="Arial", size=10, italic=True, color="888888")
    value_font = Font(name="Arial", size=11)
    note_font = Font(name="Arial", size=10, color="555555")
    note_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color=MID_GRAY),
        right=Side(style="thin", color=MID_GRAY),
        top=Side(style="thin", color=MID_GRAY),
        bottom=Side(style="thin", color=MID_GRAY),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    row = [1]  # mutable counter

    def title_row(text, height=42):
        ws.merge_cells(f"A{row[0]}:B{row[0]}")
        c = ws.cell(row=row[0], column=1, value=text)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row[0]].height = height
        row[0] += 1

    def note_row(text, bold=False, height=26):
        ws.merge_cells(f"A{row[0]}:B{row[0]}")
        c = ws.cell(row=row[0], column=1, value=text)
        c.font = Font(name="Arial", size=10, bold=bold, color=CORAL if bold else "555555")
        c.fill = note_fill; c.alignment = wrap
        ws.row_dimensions[row[0]].height = height
        row[0] += 1

    def section(text):
        ws.merge_cells(f"A{row[0]}:B{row[0]}")
        c = ws.cell(row=row[0], column=1, value=text)
        c.font = section_font; c.fill = section_fill
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row[0]].height = 30
        row[0] += 1

    def field(label, placeholder="", required=False, height=32):
        marker = " *" if required else ""
        c = ws.cell(row=row[0], column=1, value=f"{label}{marker}")
        c.font = req_label_font if required else label_font
        c.border = thin_border; c.alignment = Alignment(vertical="top")

        c = ws.cell(row=row[0], column=2, value=placeholder if placeholder else "")
        c.font = helper_font if placeholder else value_font
        c.border = thin_border; c.alignment = wrap
        ws.row_dimensions[row[0]].height = height
        row[0] += 1

    # Build the template
    title_row("CI&T WinWire — Quick Input Form")
    note_row(
        "Fill in the fields below, then attach any project documents (SOWs, decks, notes, PDFs). "
        "The skill will extract challenge/solution narratives, tech details, quotes, and page 2 "
        "content from your docs. Fields marked * are required.",
        height=48
    )
    note_row(
        "You only need to provide what the docs won't have — mainly deal numbers and a quick "
        "summary of the challenge and solution.",
        bold=True, height=30
    )

    section("PROJECT IDENTITY")
    field("Client Name", required=True)
    field("Industry", "e.g., Financial Services, Healthcare, Retail", required=True)
    field("Cloud Partner", "aws, gcp, or azure", required=True)
    field("Project Type", "e.g., Cloud Migration, Data Platform, App Modernization", required=True)
    field("Anonymize?", "Type YES to hide the client name in outputs")
    field("Anonymized Name", 'e.g., "A Leading FinTech Company" (only needed if anonymized)')

    section("DEAL METRICS (CI&T Internal)")
    note_row(
        "These numbers are typically not in project docs — please enter them manually. "
        "Services Revenue is the headline metric on the internal WinWire — without it the "
        "document will be visibly weaker.",
        height=40,
    )
    field("Services Revenue", "e.g., $3.2M — REQUIRED. Total CI&T contract value (TCV/ACV).", required=True)
    field("Incentive Funding", "e.g., $280K (MAP 2.0)", required=True)
    field("Deal Cycle", "e.g., 94 days", required=True)
    field("Incentive Program", "e.g., AWS MAP 2.0")
    field("Delivery Team Size", "e.g., 18 People")
    field("Delivery Timeline", "e.g., 9 Months")

    section("PARTNER METRICS (for partner-facing version)")
    note_row(
        "Annual Cloud Revenue is the headline metric on the partner WinWire — it's the "
        "number AWS/GCP/Azure stakeholders open the doc to see. If you don't have it yet, "
        "leave it blank and the skill will mine your attached docs; if it still can't find "
        "it, it will ask you and warn you before publishing without it.",
        height=56,
    )
    field("Annual Cloud Revenue (ACR)", "e.g., $1.2M — REQUIRED for partner version. Recurring cloud spend.", required=True)
    field("Key Business Outcome", "e.g., 40% cost reduction")
    field("Services Adopted", "e.g., 6 AWS Services")

    section("YOUR SUMMARY")
    note_row("Write a few bullet points or sentences. The skill will combine these with what it finds in your docs.")
    field("What was the challenge?", required=True, height=80)
    field("What did CI&T deliver?", required=True, height=80)

    section("LOGOS")
    note_row(
        "The CI&T and partner logos are built in — you don't need to do anything for those. "
        "For the client logo, paste a direct image URL below OR drop a PNG/SVG/JPG file into "
        "the same folder as this spreadsheet (any reasonable filename works; the skill will "
        "find it). If the skill can't download the URL, it'll ask for a new one. If no URL "
        "and no local file, the skill will ask whether you want to ship without a client logo.",
        height=70,
    )
    field("Client Logo URL", "e.g., https://example.com/acme-logo.png (direct image link)")
    field("Client Logo File", "Or the filename of a local image in this folder, e.g., acme-logo.png")

    section("OPTIONAL — OVERRIDE OR ADD")
    note_row("Only fill these in if you want to override what the skill extracts from your docs.")
    field("Preferred Title", "Leave blank to let the skill draft one from your docs")
    field("Preferred Subtitle", "Leave blank to let the skill draft one")
    field("Client Quote", "Paste a real quote if you have one — otherwise leave blank")
    field("Quote Author + Title", "e.g., Marcus Chen, CTO")

    ws.freeze_panes = "A4"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Template created: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate a simplified WinWire input template (.xlsx)")
    parser.add_argument("--output", required=True, help="Output path for the template file")
    args = parser.parse_args()
    create_template(Path(args.output))


if __name__ == "__main__":
    main()
